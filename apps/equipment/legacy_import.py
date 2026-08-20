"""
Importação assistida da planilha legada — especificação, seção 13
(fluxo D) e seção 3 ("Importação inicial da planilha atual").

Isto NÃO tenta adivinhar tudo sozinho. A aba `TOTAL EQUIPAMENTOS` da
planilha (`Estoque.Atualizado.xlsx`) usa uma taxonomia de subcategorias
(ex.: "CLIMATIZADOR COMERCIAL", variantes por BTU/voltagem) bem mais
granular do que os 9 códigos de modelo já cadastrados no catálogo atual
(seção 22 da especificação). Forçar uma correspondência automática
"errada, mas silenciosa" seria pior do que não importar — por isso cada
linha carrega uma sugestão (quando existe uma correspondência textual
razoável dentro da mesma categoria) e uma lista de "issues" quando falta
dado ou não há sugestão confiável. A decisão final de qual `EquipmentModel`
usar em cada linha é sempre humana, na tela de revisão.
"""

import difflib
import io
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook

from apps.catalog.models import Category, EquipmentModel

SHEET_NAME = "TOTAL EQUIPAMENTOS"

REQUIRED_HEADERS = {
    "categoria_pai": "Categoria (Categoria PAI)",
    "subcategoria": "Categoria  (SubCategoria)",
    "descricao_sistema": "Descrição Sistema",
    "legacy_code": "SerialNumber",
    "geracao": "Categoria:",
    "descricao": "Descrição:",
    "data_aquisicao": "Dt.Aquisicao:",
    "valor": "Valor:",
    "fornecedor": "FORNECEDOR",
}

MATCH_THRESHOLD = 0.6


class LegacyImportError(Exception):
    """Erro que impede a leitura da planilha (arquivo errado, aba faltando, etc.)."""


@dataclass
class ParsedRow:
    row_number: int
    categoria_pai: str
    subcategoria: str
    descricao_sistema: str
    legacy_code: str
    geracao: str
    descricao: str
    data_aquisicao: date | None
    valor: Decimal | None
    fornecedor: str
    suggested_model_id: int | None = None
    suggested_model_label: str = ""
    candidate_models: list[tuple[int, str]] = field(default_factory=list)  # (id, label) da mesma categoria
    issues: list[str] = field(default_factory=list)

    @property
    def has_issues(self) -> bool:
        return bool(self.issues)

    def to_session_dict(self) -> dict:
        return {
            "row_number": self.row_number,
            "categoria_pai": self.categoria_pai,
            "subcategoria": self.subcategoria,
            "descricao_sistema": self.descricao_sistema,
            "legacy_code": self.legacy_code,
            "geracao": self.geracao,
            "descricao": self.descricao,
            "data_aquisicao": self.data_aquisicao.isoformat() if self.data_aquisicao else None,
            "valor": str(self.valor) if self.valor is not None else None,
            "fornecedor": self.fornecedor,
            "suggested_model_id": self.suggested_model_id,
            "candidate_models": self.candidate_models,
            "issues": self.issues,
        }

    @classmethod
    def from_session_dict(cls, data: dict) -> "ParsedRow":
        return cls(
            row_number=data["row_number"],
            categoria_pai=data["categoria_pai"],
            subcategoria=data["subcategoria"],
            descricao_sistema=data["descricao_sistema"],
            legacy_code=data["legacy_code"],
            geracao=data["geracao"],
            descricao=data["descricao"],
            data_aquisicao=date.fromisoformat(data["data_aquisicao"]) if data["data_aquisicao"] else None,
            valor=Decimal(data["valor"]) if data["valor"] is not None else None,
            fornecedor=data["fornecedor"],
            suggested_model_id=data["suggested_model_id"],
            candidate_models=[tuple(c) for c in data["candidate_models"]],
            issues=data["issues"],
        )


def _header_index_map(header_row) -> dict[str, int]:
    found = {}
    for idx, cell in enumerate(header_row):
        if cell.value:
            found[str(cell.value).strip()] = idx
    return found


def _cell(row, index_map: dict[str, int], key: str, header_label: str):
    idx = index_map.get(header_label)
    if idx is None:
        return None
    return row[idx].value


def _to_decimal(value) -> Decimal | None:
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value))
    except InvalidOperation:
        return None


def _to_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _text(value) -> str:
    return str(value).strip() if value not in (None, "") else ""


def _best_match(targets: list[str], models: list[EquipmentModel]) -> EquipmentModel | None:
    """
    Tenta várias colunas de texto candidatas (subcategoria, descrição do
    sistema, descrição livre) e fica com a MELHOR correspondência entre
    todas elas — não apenas a primeira coluna não vazia.

    Isso importa na prática: a coluna "Descrição:" da planilha legada é
    texto livre longo (ex.: "NI48HI - AQUECEDOR EXTERNO DE AMBIENTE, A
    GÁS, TIPO TORRE, MODELONI48TPH, EM AÇO INOX, 48.000BTUS"), e comparar
    esse texto inteiro contra um nome de modelo curto ("Aquecedor Torre")
    quase sempre dá uma razão de similaridade baixa mesmo quando o texto
    contém a palavra certa — porque o SequenceMatcher pondera pelo
    tamanho combinado das strings. A subcategoria ("AQUECEDOR TORRE"),
    quando existe, é o candidato mais confiável por ser curta e já usar
    quase o mesmo vocabulário do catálogo — por isso testamos todas e
    ficamos com a melhor, em vez de travar na primeira coluna preenchida.
    """
    candidate_texts = [t for t in targets if t]
    if not candidate_texts or not models:
        return None
    best_model, best_ratio = None, 0.0
    for model in models:
        model_norm = model.name.upper()
        for target in candidate_texts:
            ratio = difflib.SequenceMatcher(None, target.upper(), model_norm).ratio()
            if ratio > best_ratio:
                best_model, best_ratio = model, ratio
    if best_ratio >= MATCH_THRESHOLD:
        return best_model
    return None


def parse_legacy_workbook(uploaded_file) -> list[ParsedRow]:
    """
    Lê a aba `TOTAL EQUIPAMENTOS` e devolve uma linha estruturada por
    equipamento, já com sugestão de modelo (quando houver correspondência
    razoável) e a lista de problemas encontrados. Nada é gravado no banco
    aqui — isso só acontece na confirmação da revisão (seção 13, fluxo D).
    """
    try:
        wb = load_workbook(io.BytesIO(uploaded_file.read()), data_only=True)
    except Exception as exc:  # openpyxl levanta várias exceções diferentes para arquivo inválido
        raise LegacyImportError(f"Não foi possível ler o arquivo como planilha Excel: {exc}") from exc

    if SHEET_NAME not in wb.sheetnames:
        raise LegacyImportError(
            f"A aba '{SHEET_NAME}' não foi encontrada. Abas disponíveis: {', '.join(wb.sheetnames)}."
        )

    ws = wb[SHEET_NAME]
    rows_iter = ws.iter_rows(min_row=1)
    header_row = next(rows_iter)
    index_map = _header_index_map(header_row)

    missing_headers = [label for label in REQUIRED_HEADERS.values() if label not in index_map]
    if missing_headers:
        raise LegacyImportError(
            "A planilha não tem as colunas esperadas: " + ", ".join(missing_headers) + ". "
            "O layout pode ter mudado desde que esta importação foi escrita."
        )

    # Categorias/modelos ativos, agrupados — usados tanto para a sugestão
    # automática quanto para popular os <select> da tela de revisão.
    models_by_category: dict[str, list[EquipmentModel]] = {}
    for category in Category.objects.filter(is_active=True):
        models_by_category[category.name.upper()] = list(
            EquipmentModel.objects.filter(category=category, is_active=True).order_by("name")
        )
    all_active_models = list(EquipmentModel.objects.filter(is_active=True).order_by("category__name", "name"))

    # Import local para evitar qualquer risco de import circular no
    # carregamento do app (legacy_import é importado por views_import,
    # que é importado no urls.py logo na subida do projeto).
    from apps.equipment.models import Equipment

    existing_legacy_codes = set(
        Equipment.objects.exclude(legacy_code="").values_list("legacy_code", flat=True)
    )

    parsed_rows: list[ParsedRow] = []

    for row_number, row in enumerate(rows_iter, start=2):
        categoria_pai = _text(_cell(row, index_map, "categoria_pai", REQUIRED_HEADERS["categoria_pai"]))
        subcategoria = _text(_cell(row, index_map, "subcategoria", REQUIRED_HEADERS["subcategoria"]))
        descricao_sistema = _text(_cell(row, index_map, "descricao_sistema", REQUIRED_HEADERS["descricao_sistema"]))
        legacy_code = _text(_cell(row, index_map, "legacy_code", REQUIRED_HEADERS["legacy_code"]))
        geracao = _text(_cell(row, index_map, "geracao", REQUIRED_HEADERS["geracao"]))
        descricao = _text(_cell(row, index_map, "descricao", REQUIRED_HEADERS["descricao"]))
        data_aquisicao = _to_date(_cell(row, index_map, "data_aquisicao", REQUIRED_HEADERS["data_aquisicao"]))
        valor = _to_decimal(_cell(row, index_map, "valor", REQUIRED_HEADERS["valor"]))
        fornecedor = _text(_cell(row, index_map, "fornecedor", REQUIRED_HEADERS["fornecedor"]))

        # Linha completamente vazia (sobra de formatação no fim da planilha) — ignorar de vez.
        if not any([categoria_pai, subcategoria, descricao_sistema, legacy_code, descricao]):
            continue

        issues: list[str] = []

        if not legacy_code:
            issues.append("Sem código de série (SerialNumber) — não é possível preservar o legacy_code.")
        elif legacy_code in existing_legacy_codes:
            issues.append(
                "Já existe um equipamento cadastrado com este código legado — provavelmente já foi "
                "importado antes. Deixe como '— não importar —' para não duplicar."
            )

        candidates = models_by_category.get(categoria_pai.upper(), [])
        category_matched = bool(categoria_pai and candidates)
        if not category_matched:
            issues.append(f"Categoria '{categoria_pai or '(vazia)'}' não corresponde a nenhuma categoria cadastrada.")
            # Sem categoria confiável para restringir a lista, mas a linha ainda
            # precisa poder ser classificada manualmente na revisão — em vez de
            # travar o <select> vazio (o que tornaria a linha impossível de
            # importar sem editar a planilha e reenviar), oferecemos todos os
            # modelos ativos como opção. A sugestão automática continua não
            # sendo tentada nesse caso (ver abaixo) porque cruzar categorias
            # erradas é justamente o tipo de acerto "por acaso" que a
            # especificação pede para evitar.
            candidates = all_active_models

        if not subcategoria and not descricao:
            issues.append("Sem subcategoria nem descrição — dado incompleto na planilha original.")

        is_duplicate = bool(legacy_code) and legacy_code in existing_legacy_codes
        suggested = (
            None
            if is_duplicate or not category_matched
            else _best_match([subcategoria, descricao_sistema, descricao], candidates)
        )
        if suggested is None and candidates and not is_duplicate and category_matched:
            issues.append("Sem correspondência automática de modelo — escolha manualmente abaixo.")

        parsed_rows.append(
            ParsedRow(
                row_number=row_number,
                categoria_pai=categoria_pai,
                subcategoria=subcategoria,
                descricao_sistema=descricao_sistema,
                legacy_code=legacy_code,
                geracao=geracao,
                descricao=descricao,
                data_aquisicao=data_aquisicao,
                valor=valor,
                fornecedor=fornecedor,
                suggested_model_id=suggested.pk if suggested else None,
                suggested_model_label=str(suggested) if suggested else "",
                candidate_models=[(m.pk, str(m)) for m in candidates],
                issues=issues,
            )
        )

    return parsed_rows
