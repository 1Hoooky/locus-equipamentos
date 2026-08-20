"""
Testes de exportação — especificação, seção 20 (critério de aceite):
"Exportação CSV/Excel reproduz corretamente patrimonio, model, status,
condition e legacy_code de todos os equipamentos ativos."
"""

import csv
import io

from django.contrib.auth import get_user_model
from django.test import TestCase
from openpyxl import load_workbook

from apps.accounts.models import Role
from apps.catalog.models import Category, EquipmentModel
from apps.equipment.services import NewEquipmentData, create_equipment

User = get_user_model()


class ExportTest(TestCase):
    def setUp(self):
        category = Category.objects.create(name="Climatizador")
        model = EquipmentModel.objects.create(category=category, name="NI23 Big Tank", code="NI23BT")
        creator = User.objects.create_user(username="cadastrador3", password="senha-forte-123")

        self.eq1 = create_equipment(
            NewEquipmentData(model_id=model.pk, created_by=creator, legacy_code="20230622201001")
        )
        self.eq2 = create_equipment(NewEquipmentData(model_id=model.pk, created_by=creator))

        User.objects.create_user(username="exporter_admin", password="senha-forte-123", role=Role.ADMIN)
        User.objects.create_user(username="exporter_consulta", password="senha-forte-123", role=Role.CONSULTA)

    def test_csv_export_contains_expected_columns_and_rows(self):
        self.client.login(username="exporter_admin", password="senha-forte-123")
        response = self.client.get("/equipamentos/exportar/?format=csv")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "text/csv")

        content = response.content.decode("utf-8")
        rows = list(csv.reader(io.StringIO(content)))
        header = rows[0]
        for expected_col in ("Patrimônio", "Código do modelo", "Status", "Condição", "Código legado"):
            self.assertIn(expected_col, header)

        patrimonio_idx = header.index("Patrimônio")
        legacy_idx = header.index("Código legado")
        exported_patrimonios = {row[patrimonio_idx] for row in rows[1:]}
        self.assertEqual(exported_patrimonios, {self.eq1.patrimonio, self.eq2.patrimonio})

        eq1_row = next(row for row in rows[1:] if row[patrimonio_idx] == self.eq1.patrimonio)
        self.assertEqual(eq1_row[legacy_idx], "20230622201001")

    def test_xlsx_export_contains_expected_rows(self):
        self.client.login(username="exporter_admin", password="senha-forte-123")
        response = self.client.get("/equipamentos/exportar/?format=xlsx")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        self.assertEqual(rows[0][0], "Patrimônio")

        exported_patrimonios = {row[0] for row in rows[1:]}
        self.assertEqual(exported_patrimonios, {self.eq1.patrimonio, self.eq2.patrimonio})

    def test_export_respects_filters(self):
        self.eq1.status = "MANUTENCAO"
        self.eq1.save(update_fields=["status"])

        self.client.login(username="exporter_admin", password="senha-forte-123")
        response = self.client.get("/equipamentos/exportar/?format=csv&status=MANUTENCAO")

        content = response.content.decode("utf-8")
        rows = list(csv.reader(io.StringIO(content)))
        self.assertEqual(len(rows), 2)  # header + 1 linha
        self.assertIn(self.eq1.patrimonio, rows[1])

    def test_consulta_role_cannot_export(self):
        self.client.login(username="exporter_consulta", password="senha-forte-123")
        response = self.client.get("/equipamentos/exportar/?format=csv")
        self.assertEqual(response.status_code, 403)
