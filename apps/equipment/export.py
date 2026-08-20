"""
Exportação CSV/Excel — especificação, seção 16/19 ("Importação e
exportação" é obrigatória; "a empresa não pode ficar refém do sistema").

As colunas exportadas reproduzem exatamente o que o critério de aceite
da Fase 1 pede (seção 20): patrimonio, model, status, condition e
legacy_code de todos os equipamentos ativos — mais alguns campos úteis
extras que não fazem mal nenhum estarem presentes.
"""

import csv

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

COLUMNS = [
    ("patrimonio", "Patrimônio"),
    ("model__code", "Código do modelo"),
    ("model__name", "Modelo"),
    ("category__name", "Categoria"),
    ("status", "Status"),
    ("condition", "Condição"),
    ("serial_number", "Serial do fabricante"),
    ("legacy_code", "Código legado"),
    ("supplier", "Fornecedor"),
    ("acquisition_date", "Data de aquisição"),
    ("acquisition_value", "Valor de aquisição"),
    ("current_location__name", "Localização atual"),
    ("current_client__company_name", "Cliente atual"),
]


def _row_values(equipment) -> list:
    status_display = equipment.get_status_display()
    condition_display = equipment.get_condition_display()
    return [
        equipment.patrimonio,
        equipment.model.code,
        equipment.model.name,
        equipment.category.name,
        status_display,
        condition_display,
        equipment.serial_number,
        equipment.legacy_code,
        equipment.supplier,
        equipment.acquisition_date.isoformat() if equipment.acquisition_date else "",
        str(equipment.acquisition_value) if equipment.acquisition_value is not None else "",
        equipment.current_location.name if equipment.current_location_id else "",
        equipment.current_client.company_name if equipment.current_client_id else "",
    ]


def export_to_csv(queryset) -> HttpResponse:
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="equipamentos-locus.csv"'

    writer = csv.writer(response)
    writer.writerow([label for _, label in COLUMNS])
    for equipment in queryset.select_related("model", "category", "current_location", "current_client"):
        writer.writerow(_row_values(equipment))

    return response


def export_to_xlsx(queryset) -> HttpResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = "Equipamentos"

    headers = [label for _, label in COLUMNS]
    ws.append(headers)
    for col_idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(header) + 2)

    for equipment in queryset.select_related("model", "category", "current_location", "current_client"):
        ws.append(_row_values(equipment))

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="equipamentos-locus.xlsx"'
    wb.save(response)
    return response
